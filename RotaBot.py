# -*- coding: utf-8 -*-
"""
Created on Wed Feb 12 14:42:46 2020

@author: benwo
"""

#Rota
import openpyxl as xl
import datetime
import random


class Cellarman():
    def __init__(self, name, dates):
        self.name = name
        self.dates = dates
        
    def get_name(self):
        return self.name
    
    def get_dates(self):
        return self.dates
        
class cellarman_getter():
    def __init__(self):
        self.cellarman = dict(zip(cellarmanClasses, [100] * len(cellarman))) #arbitrary start point utilirt = 100
        self.bar_staff = list(bar_staff)
        self.door_staff = list(door_staff)
        self.bar_recent = []
        self.door_recent = []
        
    def get_max_util_cellarman(self):
        maxUtility = 0
        maxCellarman = []
        for key,value in self.cellarman.items(): #chooses largest utility
            if value > maxUtility:
                maxUtility = value
                maxCellarman = [key]
            elif value == maxUtility:
                maxCellarman.append(key)
        return maxCellarman
        
        
    def get_cellarman(self, shift_type, date):
        maxCellarman = self.get_max_util_cellarman()    
        name = maxCellarman.pop(random.randrange(0,len(maxCellarman))) #randomly picks cellarman
        while True:
            if date in name.get_dates():
                print(name.get_name() + " cant do this date")
                self.cellarman[name] -= random.randint(2,7)
                maxCellarman = self.get_max_util_cellarman()    
                name = maxCellarman.pop(random.randrange(0,len(maxCellarman)))
            else:
                break
        print("accepted name on date" + str(date) + " is " + name.get_name())
            # run again reduce utility until somone else picked
        if shift_type == 0: # normal
            for key,value in self.cellarman.items(): #adds utility for not workng shift
                if(key!=name):
                    self.cellarman[key]+=random.randint(2,7)
                else:
                    self.cellarman[name] = 0 #resets utility to 0 so can't work again soon
        else:
            for key,value in self.cellarman.items():
                if(key!=name):
                    self.cellarman[key]+=random.randint(1,2) #adds lower utility for working busy shift
                else:
                    self.cellarman[name] = 0 #resets utility to 0 so can't work again soon
        return name.get_name()

    def get_barstaff(self):
        exit_loop = True
        while exit_loop == True:
            name = self.bar_staff.pop(random.randrange(0,len(self.bar_staff)))
            if name in self.bar_recent: #doesn't let staff work close together
                self.bar_staff.append(name)
            else:
                self.bar_recent.append(name)
                exit_loop = False                    
        if len(self.bar_staff) == 0:
            self.bar_staff = list(bar_staff)
        if len(self.bar_recent) == 30: #resets recent bar staff list
            self.bar_recent = []
        return name
    
    def get_doorstaff(self):
        exit_loop = True
        while exit_loop == True:
            name = self.door_staff.pop(random.randrange(0,len(self.door_staff)))
            if name in self.door_recent: #doesn't let staff work closer together
                self.door_staff.append(name)
            else:
                self.door_recent.append(name)
                exit_loop = False                    
        if len(self.door_staff) == 0:
            self.door_staff = list(door_staff)
        if len(self.door_recent) == 10: #resets recent door staff list
            self.door_recent = []
        return 
        

wb = xl.Workbook()
sheet = wb.active
sheet.title = "Easter Term Rota"
number_weeks = 9
cellarman=["Wolfy","Kavi","Pat","Emilibobs","SGG","Jwal","Izzy","Abi","Sam","Stonk","Adam","Liberty","Sandys","Dallas","Charlotte","Watson","Cayford","Jimbob","Redfern","Nikhil"]
cDates = [[]]
cellarmanClasses = []
week_list = ["Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]
row_titles = [" ", " ", "Cellarman", "Staff", "Staff", " ", "Quad Cellarman", "Quad Staff", " ", "Door Staff", "Door Staff", " "]
start_date = datetime.date(2020,4,26)
special_dates = [] #add days such as candlemas here
deltaday = datetime.timedelta(days=1)
number_staff = 50
number_doorstaff = 29

for name in cellarman:
    cDates = [start_date + datetime.timedelta(days=random.randint(0, number_weeks * 7)),
              start_date + datetime.timedelta(days=random.randint(0, number_weeks * 7)),
              start_date + datetime.timedelta(days=random.randint(0, number_weeks * 7))]
    print(name, cDates)
    c = Cellarman(name, cDates)
    cellarmanClasses.append(c)
    

#read bar staff from txt file
bar_staff = []
file_staff = open("barstaff.txt","r")
file_staff1 = file_staff.readlines()
for i in range(number_staff):
    bar_staff.append(file_staff1[4*i])

#read door staff from txt file
door_staff = []
file_door = open("doorstaff.txt","r")
file_door1 = file_door.readlines()
for i in range(number_doorstaff):
    door_staff.append(file_door1[4*i])

cg = cellarman_getter()

#make it look pretty
for i in range(3,122):
    sheet.row_dimensions[i].height = 30
for i in range(8):
    sheet.column_dimensions[xl.cell.cell.get_column_letter(i+1)].width = 20
for i in range(number_weeks):
    for day in range(7):
        if i == number_weeks-1 and day == 6: #last saturday of term
            break
        else:
            current_cell1 = sheet.cell(row = (12*i)+3,column = day+2)
            current_cell1.fill = xl.styles.PatternFill(fgColor="85C1E9", fill_type = "solid")
            current_cell2 = sheet.cell(row = (12*i)+4,column = day+2)
            current_cell2.fill = xl.styles.PatternFill(fgColor="EC7063", fill_type = "solid")

# Days of week loop 
#Week
for i in range(number_weeks):
    #days
    for day in range(7):
        if i == number_weeks-1 and day == 6: #last sat of term
            break
        else:
            current_cell = sheet.cell(row = (12*i)+3,column = day+2)
            current_cell.value = week_list[day]
        

#Row heading loop
for i in range(number_weeks):
    for j in range(12):
        current_cell = sheet.cell(row=3+j+(12*i),column=1)
        current_cell.value = row_titles[j]
        current_cell.fill = xl.styles.PatternFill(fgColor="2ECC71",fill_type="solid")
     
#Date loop
current_date = start_date
for i in range(number_weeks):
    for day in range (7):
        if i == number_weeks-1 and day == 6: #last sat of term
            break
        else:
            current_cell = sheet.cell(row = (12*i)+4, column = day + 2)
            current_cell.value = str(current_date)
            current_date = current_date + deltaday
        
#Bar Staff 
for i in range(number_weeks):
    for day in range(7):
        if i == number_weeks-1 and day == 6: #last sat of term
            break
        else:  
            for k in range(2):
                current_cell = sheet.cell(row = (12*i)+6+k, column = day + 2)
                current_cell.value = cg.get_barstaff()
                if (day == 3) or (day == 5) or (day == 6) or (i == number_weeks -1 and day == 4): #busy shift and megaformal
                    current_cell3 = sheet.cell(row = (12*i)+10, column = day + 2)
                    current_cell3.value = cg.get_barstaff()
                    if i == number_weeks-1 and day == 5: #last day of term
                        current_cell3.value = "Barcom"
                if (i == 0 and day == 0) or (i == number_weeks-1 and day == 5):
                    current_cell.value = "Barcom"


#Door Staff
for i in range(number_weeks):
    for day in range(7):
        if (day == 3) or (day == 5) or (day == 6):
            if i == number_weeks-1 and day == 6: #last sat of term
                break
            elif i == number_weeks-1 and day == 5: #last day of term
                break
            elif i == number_weeks-1 and day == 4: #megaformal
                for k in range(2):
                    current_cell = sheet.cell(row = (12*i)+12+k, column = day + 2)
                    current_cell.value = cg.get_doorstaff()
            else:
                for k in range(2): #2 staff
                    current_cell = sheet.cell(row = (12*i)+12+k, column = day + 2)
                    current_cell.value = cg.get_doorstaff()
                    if (i == number_weeks-1) and day == 5: #last day of term
                        current_cell.value = "Barcom"
                    
                
#Cellarman 
current_date = start_date - deltaday
for i in range(number_weeks):
    for day in range(7):
        if (i == number_weeks-1) and day == 6:
            break
        else:
            current_cell = sheet.cell(row = (12*i)+5, column = day + 2)
            current_date = current_date + deltaday
            current_cell.value = cg.get_cellarman(0,current_date)
            if (i == 0 and day == 0) or (i == number_weeks-1 and day == 5): #first and last day or term
                current_cell.value = "Barcom"
            
            if (day == 3) or (day == 5) or (day == 6) or (i == number_weeks -1 and day == 4): #busy shifts and megaformal
              if (i == number_weeks-1 and day == 5): #last day of term
                  current_cell = sheet.cell(row = (12*i)+9, column = day + 2)
                  current_cell.value = "Barcom"
                  break
              while True:
                  potential_cellarman = cg.get_cellarman(1,current_date)
                  if potential_cellarman != current_cell.value: #stops quad and main being same
                      current_cell = sheet.cell(row = (12*i)+9, column = day + 2)
                      current_cell.value = potential_cellarman
                      break

wb.save(r"C:\Users\benwo\OneDrive\Documents\Chad's Year 3\Bar Manager\rota.xlsx")
